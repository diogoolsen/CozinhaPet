# conda create --name cozinhaPet
# conda install -c anaconda openpyxl


from openpyxl import Workbook
from openpyxl import load_workbook

# from openpyxl import Workbook
arquivo_excel = Workbook()

# caminho = '/caminho/até/o/seu/arquivo.xlsx'
# arquivo_excel = load_workbook(caminho)

wb = load_workbook('teste.xlsx')

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['B2'] = 42
ws['B4'] = 42
ws['D4'] = 42

# Rows can also be appended
# ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")
